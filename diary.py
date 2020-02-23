#!/usr/bin/python3
#!/usr/bin/env python3

import os
import re
import sys
import time as TIME
import getopt
import socket
from ftplib import FTP
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import configparser
from datetime import date,datetime,time

import readline


## configuare file
def get_server_mode( config_file_name="config.ini" ):
    if config_file_name == "":
        return None
    else :
        if not config_file_name.endswith(".ini") :
            return None
    config_file_path  = sys.path[0] +'/'+config_file_name
    cfg_file = configparser.ConfigParser()
    cfg_file.read( config_file_path)
    if "FTP Server" in cfg_file :
        server_online_str = cfg_file["FTP Server"]["online"]
    else :
        server_online_str= "n"
    return server_online_str

def get_server_user_pwd( config_file_name="config.ini" ):
    if config_file_name == "":
        return None
    else :
        if not config_file_name.endswith(".ini") :
            return None

    config_file_path  = sys.path[0] +'/'+config_file_name
    cfg_file = configparser.ConfigParser()
    cfg_file.read( config_file_path)
    if "FTP Server" in cfg_file :
        server_user_str = cfg_file["FTP Server"]["userName"]
        server_pass_str = cfg_file["FTP Server"]["password"]
    else :
        server_user_str = "user"
        server_pass_str = "123"
    return server_user_str,server_pass_str

def get_server_address( config_file_name="config.ini" ):
    if config_file_name == "":
        return None
    else :
        if not config_file_name.endswith(".ini") :
            return None
    config_file_path  = sys.path[0] +'/'+config_file_name
    cfg_file = configparser.ConfigParser()
    cfg_file.read( config_file_path)

    if "FTP Server" in cfg_file :
        server_ip_str = cfg_file["FTP Server"]["serverIP"]
        server_port_str = cfg_file["FTP Server"]["serverPort"]
    else :
        server_ip_str = "192.168.1.105"
        server_port_str = "21"
    return server_ip_str,server_port_str


def get_auto_upload( config_file_name="config.ini" ):
    if config_file_name == "":
        print ("配置文件名为空")
        return None
    else :
        if not config_file_name.endswith(".ini") :
            print ("无效的配置文件名")
            return None
    config_file_path  = sys.path[0] +'/'+config_file_name
    if not os.access( config_file_path , os.F_OK ):
        print ("配置文件访问失败")
        return None
    else :
        cfg_file = configparser.ConfigParser()
        cfg_file.read( config_file_path)

        if "DEFAULT" in cfg_file :
            auto_upload_int = cfg_file["DEFAULT"].getint("auto_upload")
        else :
            auto_upload_int = 1
    return auto_upload_int

#### date and time
def get_date():
    return datetime.now().date().strftime("%Y-%m-%d")

def get_week():
    cur_week_in_year_tuple = datetime.now().isocalendar()
    return cur_week_in_year_tuple[1]

def get_day_in_week():
    cur_week_in_year_tuple = datetime.now().isocalendar()
    return cur_week_in_year_tuple[2]

def get_time():
    return datetime.now().time().strftime("%H:%M:%S")

def get_sheet( file,wb_obj , title_str ):
    try:
        sheet_list = wb_obj.get_sheet_names()
        sheet_index = sheet_list.index( title_str )
    except ValueError :
        temp_sheet_index = sheet_list.index( "0" )
        sheet_obj = wb_obj.copy_worksheet ( wb_obj.worksheets[ temp_sheet_index ]  )
        sheet_obj.title = title_str
        wb_obj.save( file )
    else :
        sheet_obj = wb_obj.worksheets[ sheet_index ]
    finally:
        return sheet_obj

def ftp_upload( remote_file , local_file , ip,  userName, password ):
    try:
        ftp = FTP(host=ip , user=userName , passwd=password)
    except (socket.error, socket.gaierror):
        print ("异常：服务器连接失败")
        return None
    
    ftp.login( user=userName , passwd=password)
    print ( "登陆成功:  %s" % ftp.getwelcome())
    # 匿名登陆，文件写入 ftp-anonymous/ 
    # ftp = FTP(host="192.168.1.105")
    # ftp.login()
    with open( local_file,"rb") as fd :
        ftp.storbinary('STOR ' + remote_file , fd ,1024 )
    ftp.quit()

def show_week( file,week_int ):
    wb = get_workbook(  file )
    sheet = get_sheet( file , wb , str(week_int) )
    year_week = str( "%s-%i-"%(datetime.now().date().strftime("%Y"),week_int) )
    day_in_week_str = ["周一","周二","周三","周四","周五","周六","周日"]
   
    print( " %s "% str( "< %s 年  第 %i 周 >"%(datetime.now().date().strftime("%Y"),week_int) ))
    for i in range(1,8):
        date_str = TIME.strftime("%Y-%m-%d",TIME.strptime( year_week+str(i) ,'%Y-%W-%u'))
        content = sheet.cell( i+1 ,2).value
        if content != None :
            print( "----------------------")
            print( "|  %s   %s |"%  (day_in_week_str[i-1],date_str))
            print( "----------------------")
            print( "%s"% content )

## general operation
def commit( file,content_str ):
    if content_str != "":
        week_int = get_week()
        day_in_week_int = get_day_in_week()
        print ("%s      第%d周 第%d天" % ( get_date(), week_int ,day_in_week_int ))

        wb = get_workbook(  file )
        sheet = get_sheet( file , wb ,str(week_int) )
        privous_str = sheet.cell( day_in_week_int+1 ,2).value
        sheet.cell( day_in_week_int+1 ,2).alignment = Alignment( horizontal="left",vertical="top")
        
        time_str = get_time()
        if privous_str == None :
            # 空白cell
            sheet.cell( day_in_week_int+1 ,2).value = "["+ time_str+"] "+ content_str
        else :
            sheet.cell( day_in_week_int+1 ,2).value = privous_str +"\n["+ time_str+"]  "+ content_str
        wb.save( file )
    else :
        print (" 输入内容为空，不写入任何内容")


## 测试: test_diary.test_get_workbook()
def get_workbook( fileName = "yangj_log.xlsx" ):
    if fileName == "":
        return None
    else :
        if not fileName.endswith(".xlsx")  :
            return None

    if not os.access( fileName , os.F_OK ):
        # 新建文件及 wb，#0 sheet，向其中填充模板
        wb = openpyxl.Workbook()
        sheet = wb .create_sheet("0") 
        day_in_week_str = ["周一","周二","周三","周四","周五","周六","周日"]
        sheet.cell( 1 ,1).value = "日期"
        sheet.cell( 1 ,2).value = "内容"
        for i in range(0,7):
            print ( i,day_in_week_str[i] )
            sheet.cell( i+2 ,1).value = day_in_week_str[i]
        wb.save(fileName)
        print (" File not found ,then create new workbook")
        return wb
    return openpyxl.load_workbook( fileName )

def get_reocrd_file_name( config_file_name="config.ini" ):
    if config_file_name == "":
        return None
    else :
        if not config_file_name.endswith(".ini") :
            return None
    config_file_path  = sys.path[0] +'/'+config_file_name
    cfg_file = configparser.ConfigParser()
    cfg_file.read( config_file_path)
    if "DEFAULT" in cfg_file :
        local_file_name  = cfg_file["DEFAULT"]["local_file_name"]
    else :
        local_file_name="local_diary.xlsx"
    return local_file_name

## 创建配置文件
def create_config( config_file_path="config.ini"):
    config = configparser.ConfigParser()
    config["FTP Server"] = {}

    server_ip_str = input(" 服务器IP : ")
    if not re.match(  r'^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$' ,server_ip_str):
        print ( "IP address invaild ")
        exit()
    config["FTP Server"]["serverIP"] = server_ip_str

    server_port_str = input(" 服务器Port : ")
    if not re.match(  r'^(?:[0-9]{1,5})$' ,server_port_str):
        print ( "Port invaild ")
        exit()
    config["FTP Server"]["serverPort"] = server_port_str

    server_online_str = input(" 网络支持 (y/n): ")
    if server_online_str != "y" :
        server_online_str = "n"
    config["FTP Server"]["online"] = server_online_str

    server_user_str = input(" 用户名: ")
    config["FTP Server"]["userName"] = server_user_str

    server_pass_str = input(" 用户密码: ")
    config["FTP Server"]["password"] = server_pass_str

    config["DEFAULT"] = {}
    local_file_name_str = input(" 日志文件(*.xlsx): ")
    config["DEFAULT"]["local_file_name"] = local_file_name_str
    config["DEFAULT"]["auto_upload"] = "1" ## 默认1，未来可能取消自动推送功能，改用 push 命令推送

    #TODO:写异常需要处理
    with open(  config_file_path ,"w") as fd :
        config.write( fd ) 
    return 0

def check_environment( ):
    config_file_path  = sys.path[0] +'/config.ini'
    if not os.access( config_file_path , os.F_OK ):
        print (" 首次使用,配置服务器参数：")
        create_config( config_file_path)
    
    record_file_name = get_reocrd_file_name()
    record_file_path  = sys.path[0] +'/'+record_file_name
    if not os.access( record_file_path , os.F_OK ):
        get_workbook(  record_file_name )
    return 0

############################
# @berif 打印 help 信息
def usage( ):
    print(' Usage: diary <cmd> [opt]')
    print(' cmd:')
    print('     commit  ：提交一条记录，Enter 结束输入')
    print('     show  [ week ] : 显示第 week 周的内容')
    print('     push    : 向 FTP 服务器推送本地记录文件')
    print(' Diary v0.0.1  2020/2/3 ( leoyang20102013@163.com )')

#################
def main():
    if sys.argv.__len__() <= 1:
        usage()
        exit()

    if check_environment() < 0 :
        exit()

        

    local_file_path  = sys.path[0] +'/' + get_reocrd_file_name()
    if sys.argv[1] == "commit" :
        commit( local_file_path,input("[随手记]")  )
        exit()
    elif sys.argv[1] == "show" :
        week_int = get_week()
        if (sys.argv.__len__() > 2) and (sys.argv[2].isdigit()) :
            w = int(sys.argv[2])
            if w > week_int :
                w = week_int
        else :
            w = week_int
        show_week( local_file_path,w )
        exit()
    elif (sys.argv[1] == "push") and (server_online_str == "y") :
        ret = ftp_upload( local_file_name , local_file_path ,server_ip_str, server_user_str, server_pass_str )
        if ret == None :
            print ("确认服务器IP配置正确或服务器已经启动")
        exit()
    elif sys.argv[1] == "help" :
        usage( )
        exit()
    else:
        exit()
    return 

#########
if __name__ == "__main__":
    main()